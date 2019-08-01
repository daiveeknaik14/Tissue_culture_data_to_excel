from tkinter import *
import openpyxl as py
import os

ws = py.load_workbook("example1.xlsx")
sheet = ws["Sheet1"]

sheet.cell(row=1, column=1).value = "Number"
sheet.cell(row=1, column=2).value = "Invoice no"
sheet.cell(row=1, column=3).value = "Date"
sheet.cell(row=1, column=4).value = "Name of Farmer"
sheet.cell(row=1, column=5).value = "Type"
sheet.cell(row=1, column=6).value = "Variety"
sheet.cell(row=1, column=7).value = "Village"
sheet.cell(row=1, column=8).value = "Dealer"
sheet.cell(row=1, column=9).value = "Plants"
sheet.cell(row=1, column=10).value = "Plants Return"
sheet.cell(row=1, column=11).value = "Rate"
sheet.cell(row=1, column=12).value = "Total Payment"
sheet.cell(row=1, column=13).value = "Vehicle"
sheet.cell(row=1, column=14).value = "Name"
sheet.cell(row=1, column=15).value = "Kilometers"
sheet.cell(row=1, column=16).value = "Upad/Payment"
sheet.cell(row=1, column=17).value = "Cheq no"
sheet.cell(row=1, column=18).value = "Return"
sheet.cell(row=1, column=19).value = "Remark"

ws.save("example1.xlsx")

def delete_login_success():
    root.destroy()

def login_sucess():
    global root
    root = Toplevel(login_screen)
    def insert():
        displaylabel1 = Label(root, text="Added Successfully")
        displaylabel1.grid(row=14, column=4)
        current_row = sheet.max_row
        current_column = sheet.max_column
        sheet.cell(row=current_row + 1, column=1).value = numbertext.get()
        numbertext.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=2).value = invoicetext.get()
        invoicetext.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=3).value = datetext1.get()
        datetext1.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=4).value = nametext1.get()
        nametext1.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=5).value = typetext.get()
        typetext.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=6).value = varietytext1.get()
        varietytext1.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=7).value = villagetext.get()
        villagetext.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=8).value = dealertext.get()
        dealertext.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=9).value = plantstext.get()
        sheet.cell(row=current_row + 1, column=10).value = returntext1.get()
        sheet.cell(row=current_row + 1, column=11).value = ratetext.get()
        a = float(plantstext.get())
        b = float(returntext1.get())
        c = float(ratetext.get())
        returntext1.delete(0, 'end')
        ratetext.delete(0, 'end')
        plantstext.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=12).value = (a-b)*c
        sheet.cell(row=current_row + 1, column=13).value = vehicletext.get()
        vehicletext.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=14).value = nametext2.get()
        nametext2.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=15).value = kmtext.get()
        kmtext.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=16).value = paymenttext.get()
        paymenttext.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=17).value = cheqtext.get()
        cheqtext.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=18).value = return2text.get()
        return2text.delete(0, 'end')
        sheet.cell(row=current_row + 1, column=19).value = remarktext.get()
        remarktext.delete(0, 'end')

        ws.save("example1.xlsx")

    def click1():
        displaylabel2 = Label(root, text="Updated Successfully")
        displaylabel2.grid(row=18, column=5)
        t = costupdatetext.get()
        t1 = searchnametext.get()
        t5 = searchdatetext.get()
        for row in range(2, sheet.max_row + 1):
            c2 = sheet.cell(row=row, column=4).value
            c3 = sheet.cell(row=row, column=3).value
            if c3 == t5:
                sheet.cell(row=row, column=12).value = t
        ws.save("example1.xlsx")
        costupdatetext.delete(0, 'end')

    def click():

        t1 = searchnametext.get()
        t5 = searchdatetext.get()
        for row in range(2, sheet.max_row + 1):
            c2 = sheet.cell(row=row, column=4).value
            c3 = sheet.cell(row=row, column=3).value
            if c2 == t1 and c3 == t5:
                t2 = sheet.cell(row=row, column=9).value
                t3 = sheet.cell(row=row, column=10).value
                t4 = sheet.cell(row=row, column=12).value

        line14 = Label(root, text='')
        line14.grid(row=18, column=1)

        display1 = Label(root, text="Plants :")
        display1.grid(row=19, column=1)

        display2 = Label(root, text=t2)
        display2.grid(row=19, column=2)

        display3 = Label(root, text="Total :")
        display3.grid(row=19, column=3)

        display4 = Label(root, text=t3)
        display4.grid(row=19, column=4)

        display3 = Label(root, text="Payment :")
        display3.grid(row=19, column=5)

        display6 = Label(root, text=t4)
        display6.grid(row=19, column=6)

        searchnametext.delete(0, 'end')

    def on_keyrelease(event):

        # get text from entry
        value = event.widget.get()
        value = value.strip().lower()

        # get data from test_list
        if value == '':
            data = list
        else:
            data = []
            for item in list:
                if value in item.lower():
                    data.append(item)

        # update data in listbox
        listbox_update(data)

    def listbox_update(data):
        # delete previous data
        listbox.delete(0, 'end')

        # sorting data
        data = sorted(data, key=str.lower)

        # put new data
        for item in data: listbox.insert('end', item)

    def on_select(event):
        # display element selected on list
        print('(event) previous:', event.widget.get('active'))
        print('(event)  current:', event.widget.get(event.widget.curselection()))
        print('---')
    list = []
    m_row = sheet.max_row
    for i in range(2, m_row + 1):
        cell_obj = sheet.cell(row=i, column=3)
        list.append(cell_obj.value)

    root = Tk()
    root.title('Delivery')
    root.geometry('800x800')

    heading1 = Label(root, text='Delivery Details')
    heading1.grid(row=1, column=1)
    heading1.config(font=100)

    line1 = Label(root, text='------------------------')
    line1.grid(row=2, column=1)

    number = Label(root, text='Number')
    number.grid(row=3, column=1)
    numbertext = Entry(root)
    numbertext.grid(row=3, column=2)

    line2 = Label(root, text='------------------------')
    line2.grid(row=2, column=2)

    line3 = Label(root, text='------------------------')
    line3.grid(row=2, column=3)

    line4 = Label(root, text='------------------------')
    line4.grid(row=2, column=4)

    line5 = Label(root, text='------------------------')
    line5.grid(row=2, column=5)

    line6 = Label(root, text='------------------------')
    line6.grid(row=2, column=6)

    name1 = Label(root, text='Name of Farmer')
    name1.grid(row=3, column=3)
    nametext1 = Entry(root)
    nametext1.grid(row=3, column=4)

    variety = Label(root, text='Variety')
    variety.grid(row=3, column=5)
    varietytext1 = Entry(root)
    varietytext1.grid(row=3, column=6)

    line7 = Label(root, text='')
    line7.grid(row=4, column=1)

    date1 = Label(root, text='Date')
    date1.grid(row=5, column=1)
    datetext1 = Entry(root)
    datetext1.grid(row=5, column=2)

    type = Label(root, text='Type')
    type.grid(row=5, column=3)
    typetext = Entry(root)
    typetext.grid(row=5, column=4)

    village = Label(root, text='Village')
    village.grid(row=5, column=5)
    villagetext = Entry(root)
    villagetext.grid(row=5, column=6)

    line8 = Label(root, text='')
    line8.grid(row=6, column=1)

    dealer = Label(root, text='Dealer')
    dealer.grid(row=7, column=1)
    dealertext = Entry(root)
    dealertext.grid(row=7, column=2)

    return1 = Label(root, text='Return Plants')
    return1.grid(row=7, column=3)
    returntext1 = Entry(root)
    returntext1.grid(row=7, column=4)

    vehicle = Label(root, text='Vehicle')
    vehicle.grid(row=7, column=5)
    vehicletext = Entry(root)
    vehicletext.grid(row=7, column=6)

    line9 = Label(root, text='')
    line9.grid(row=8, column=1)

    plants = Label(root, text='Plants')
    plants.grid(row=9, column=1)
    plantstext = Entry(root)
    plantstext.grid(row=9, column=2)

    rate = Label(root, text='Rate')
    rate.grid(row=9, column=3)
    ratetext = Entry(root)
    ratetext.grid(row=9, column=4)

    name2 = Label(root, text='Driver Name')
    name2.grid(row=9, column=5)
    nametext2 = Entry(root)
    nametext2.grid(row=9, column=6)

    line10 = Label(root, text='')
    line10.grid(row=10, column=1)

    km = Label(root, text='Kilometers')
    km.grid(row=11, column=1)
    kmtext = Entry(root)
    kmtext.grid(row=11, column=2)

    payment = Label(root, text='Upad/Payment')
    payment.grid(row=11, column=3)
    paymenttext = Entry(root)
    paymenttext.grid(row=11, column=4)

    cheq = Label(root, text='Cheq no.')
    cheq.grid(row=11, column=5)
    cheqtext = Entry(root)
    cheqtext.grid(row=11, column=6)

    line10 = Label(root, text='')
    line10.grid(row=12, column=1)

    return2 = Label(root, text='Return')
    return2.grid(row=13, column=1)
    return2text = Entry(root)
    return2text.grid(row=13, column=2)

    remark = Label(root, text='Remark')
    remark.grid(row=13, column=3)
    remarktext = Entry(root)
    remarktext.grid(row=13, column=4)

    invoice = Label(root, text='Invoice no')
    invoice.grid(row=13, column=5)
    invoicetext = Entry(root)
    invoicetext.grid(row=13, column=6)

    add = Button(root, text='ADD', width=15, command=insert)
    add.grid(row=14, column=6)

    heading2 = Label(root, text='Search and Update')
    heading2.grid(row=15, column=1)
    heading2.config(font=100)

    line12 = Label(root, text='------------------------')
    line12.grid(row=16, column=1)

    searchname = Label(root, text='Name Search')
    searchname.grid(row=17, column=1)
    searchnametext = Entry(root)
    searchnametext.grid(row=17, column=2)

    searchdate = Label(root, text='DateSearch')
    searchdate.grid(row=20, column=1)
    searchdatetext = Entry(root)
    searchdatetext.grid(row=20, column=2)
    searchdatetext.bind('<KeyRelease>', on_keyrelease)

    line13 = Label(root, text='')
    line13.grid(row=18, column=1)

    line15 = Label(root, text='')
    line15.grid(row=20, column=1)

    listbox = Listbox(root)
    listbox.grid(row=21, column=2)
    listbox.bind('<<ListboxSelect>>', on_select)
    listbox_update(list)

    searchbutton = Button(root, text='Search', command=click, font=12)
    searchbutton.grid(row=17, column=3)

    costupdate = Label(root, text="Amount Update")
    costupdate.grid(row=17, column=4)
    costupdatetext = Entry(root)
    costupdatetext.grid(row=17, column=5)

    updatebutton = Button(root, text='Update', command=click1, font=12)
    updatebutton.grid(row=17, column=6)


def delete_password_not_recognised():
    password_not_recog_screen.destroy()

def password_not_recognised():
    global password_not_recog_screen
    password_not_recog_screen = Toplevel(login_screen)
    password_not_recog_screen.title("Success")
    password_not_recog_screen.geometry("150x100")
    passwordtext = Label(password_not_recog_screen, text="Invalid Password ")
    passwordtext.grid(row=2, column=2)
    passwordbutton = Button(password_not_recog_screen, text="OK", command=delete_password_not_recognised)
    passwordbutton.grid(row=4, column=2)

def delete_user_not_found_screen():
    user_not_found_screen.destroy()

def user_not_found():
    global user_not_found_screen
    user_not_found_screen = Toplevel(login_screen)
    user_not_found_screen.title("Success")
    user_not_found_screen.geometry("150x100")
    usertext = Label(user_not_found_screen, text="User Not Found")
    usertext.grid(row=2, column=2)
    userbutton = Button(user_not_found_screen, text="OK", command=delete_user_not_found_screen)
    userbutton.grid(row=4, column=2)

def login_verify():
    username1 = username_verify.get()
    password1 = password_verify.get()
    username_login_entry.delete(0, END)
    password_login_entry.delete(0, END)

    list_of_files = os.listdir()
    if username1 in list_of_files:
        file1 = open(username1, "r")
        verify = file1.read().splitlines()
        if password1 in verify:
            login_sucess()

        else:
            password_not_recognised()

    else:
        user_not_found()

def login():
    global login_screen
    login_screen = Toplevel(main_screen)
    login_screen.title("Login")
    login_screen.geometry("500x250")
    logintext = Label(login_screen, text="Please enter details below to login")
    logintext.grid(row=1, column=1)

    global username_verify
    global password_verify

    username_verify = StringVar()
    password_verify = StringVar()

    global username_login_entry
    global password_login_entry

    loginuser = Label(login_screen, text="Username * ")
    username_login_entry = Entry(login_screen, textvariable=username_verify)
    loginuser.grid(row=2, column=2)
    username_login_entry.grid(row=3,column=3)

    loginpassword = Label(login_screen, text="Password * ")
    password_login_entry = Entry(login_screen, textvariable=password_verify, show='*')
    loginpassword.grid(row=4, column=2)
    password_login_entry.grid(row=5, column=3)

    loginbutton = Button(login_screen, text="Login", width=10, height=1, command=login_verify)
    loginbutton.grid(row=7, column=4)

def register_user():
    username_info = username.get()
    password_info = password.get()

    file = open(username_info, "w")
    file.write(username_info + "\n")
    file.write(password_info)
    file.close()

    username_entry.delete(0, END)
    password_entry.delete(0, END)

    usertext = Label(register_screen, text="Registration Success", fg="green", font=("calibri", 11))
    usertext.grid(row=2, column=2)

def register():
    global register_screen
    register_screen = Toplevel(main_screen)
    register_screen.title("Register")
    register_screen.geometry("500x250")

    global username
    global password
    global username_entry
    global password_entry
    username = StringVar()
    password = StringVar()

    registertext = Label(register_screen, text="Please enter details below", bg="blue")
    registertext.grid(row=1, column=1)
    username_lable = Label(register_screen, text="Username * ")
    username_lable.grid(row=3, column=2)
    username_entry = Entry(register_screen, textvariable=username)
    username_entry.grid(row=2, column=3)
    password_lable = Label(register_screen, text="Password * ")
    password_lable.grid(row=4, column=2)
    password_entry = Entry(register_screen, textvariable=password, show='*')
    password_entry.grid(row=4, column=3)
    registerbutton = Button(register_screen, text="Register", width=10, height=1, bg="blue", command=register_user)
    registerbutton.grid(row=5, column=3)

def main_account_screen():
    global main_screen
    main_screen = Tk()
    main_screen.geometry("250x250")
    main_screen.title("Account Login")
    maintext = Label(main_screen, text="Select Your Choice")
    maintext.grid(row=1, column=2)
    mainbutton1 = Button(main_screen, text="Login", width=15, command=login)
    mainbutton1.grid(row=3, column=2)
    mainbutton2 = Button(main_screen, text="Register", width=15, command=register)
    mainbutton2.grid(row=5, column=2)

    main_screen.mainloop()


main_account_screen()